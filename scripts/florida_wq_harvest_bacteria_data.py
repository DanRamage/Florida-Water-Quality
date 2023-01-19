import sys
sys.path.append('../commonfiles/python')
import os
import time
import logging.config
import json
from openpyxl import load_workbook
import poplib
import email
import optparse
if sys.version_info[0] < 3:
  import ConfigParser
else:
  import configparser as ConfigParser
import json
import datetime

from florida_wq_data import florida_sample_sites
from advisory import florida_wq_advisory

def contains(list, filter):
  for x in list:
    if filter(x):
      return True
  return False

def check_email_for_update(config_file):
  file_list = []
  logger = logging.getLogger('check_email_for_update_logger')
  logger.debug("Starting check_email_for_update")
  try:
    email_ini_config_filename = config_file.get("email_settings", "settings_ini")
    email_ini_config_file = ConfigParser.RawConfigParser()
    email_ini_config_file.read(email_ini_config_filename)

    email_host = email_ini_config_file.get("sarasota_email_settings", "host")
    email_user = email_ini_config_file.get("sarasota_email_settings", "user")
    email_password = email_ini_config_file.get("sarasota_email_settings", "password")
    destination_directory = config_file.get("worksheet_settings", "destination_directory")
  except ConfigParser.Error as e:
    if logger:
      logger.exception(e)
  try:
    if logger:
      logger.info("Attempting to connect to email server.")

    pop3_obj = poplib.POP3_SSL(email_host, 995)
    pop3_obj.user(email_user)
    pop3_obj.pass_(email_password)

    if logger:
      logger.info("Successfully connected to email server.")
  except (poplib.error_proto, Exception) as e:
    if logger:
      logger.exception(e)
  else:
    #messages = {}
    #for i in range(1, len(pop3_obj.list()[1]) + 1):
    ##  resp, message, byteCnt = pop3_obj.retr(i)
    #  messages[i] = message
    # messages processing
    try:
      emails, total_bytes = pop3_obj.stat()
      for i in range(emails):
          # return in format: (response, ['line', ...], octets)
          response = pop3_obj.retr(i+1)
          raw_message = response[1]

          str_message = email.message_from_string((b"\n".join(raw_message)).decode('utf-8'))

          # save attach
          for part in str_message.walk():
            if logger:
              logger.debug("Content type: %s" % (part.get_content_type()))

            if part.get_content_maintype() == 'multipart':
              continue

            if part.get('Content-Disposition') is None:
              if logger:
                logger.debug("No content disposition")
              continue

            filename = part.get_filename()
            if filename.find('xlsx') != -1:
              if logger:
                logger.debug("Attached filename: %s" % (filename))

              saved_file_name = os.path.join(destination_directory, filename)
              with open(saved_file_name, 'wb') as out_file:
                out_file.write(part.get_payload(decode=1))
                out_file.close()
                file_list.append(saved_file_name)
              #parse_sheet_data(saved_file_name, use_logging)
    except Exception as e:
      logger.exception(e)

    #pop3_obj.quit()

  if logger:
    logger.debug("Finished check_email_for_update")
  return file_list

def parse_sheet_data(xl_file_name):
  logger = logging.getLogger('check_email_for_update_logger')
  logger.debug("Starting parse_sheet_data, parsing file: %s" % (xl_file_name))


  sample_data = {}
  wb = load_workbook(filename = xl_file_name)
  bacteria_data_sheet = wb['SCHD Bact Results']
  date_column = 'A'
  time_column = 'B'
  station_column = 'D'
  entero_column = 'H'
  sample_dates = []
  for row_num in range(2,18):
    cell_name = "%s%d" % (station_column, row_num)
    cell_value = "%s%d" % (entero_column, row_num)
    cell_date = "%s%d" % (date_column, row_num)
    cell_time = "%s%d" % (time_column, row_num)
    try:
      if isinstance(bacteria_data_sheet[cell_time].value, datetime.time):
        sample_date_time = datetime.datetime.combine(bacteria_data_sheet[cell_date].value, bacteria_data_sheet[cell_time].value)
      else:
        sample_date_time = bacteria_data_sheet[cell_date].value
    except Exception as e:
      sample_date_time = bacteria_data_sheet[cell_date].value
      logger.exception(e)
    site_name = bacteria_data_sheet[cell_name].value.strip().lower()

    if site_name not in sample_data:
      sample_data[site_name] = []
    sample_data[site_name].append({
      'site_name': site_name,
      'sample_date': sample_date_time,
      'value': bacteria_data_sheet[cell_value].value
    })

    #sample_data[site_name].append({
    #  'sample_date': sample_date_time,
    #  'value': bacteria_data_sheet[cell_value].value
    #})
    if bacteria_data_sheet[cell_date].value not in sample_dates:
      sample_dates.append(bacteria_data_sheet[cell_date].value)
    #if sample_date is None:
    #  sample_date = bacteria_data_sheet[cell_date].value
  if logger:
    logger.debug("Finished parse_sheet_data")

  return sample_data, sample_dates

def parse_suncoast_sheet_data(xl_file_name):
  logger = logging.getLogger()
  logger.debug("Starting parse_suncoast_sheet_data, parsing file: %s" % (xl_file_name))

  sample_data = {}
  wb = load_workbook(filename = xl_file_name)

  bacteria_data_sheet = wb['Sheet1']
  station_column = 'A'
  sample_dates = []
  #This sheet is organized with station name in first column, then the next columns are the dates of the samples.
  #The first row has the dates.
  for row_ndx,col in enumerate(bacteria_data_sheet.iter_rows(max_row=13)):
    # This will have the dates in the column headers.
    if row_ndx == 0:
      for cell_ndx,cell in enumerate(col):
        #First cell is the Sample SIte header on the first row, so we ignore it.
        if cell_ndx > 0:
          #Date format at the moment is "11.16.22".
          date = cell.value
          try:
            sample_datetime = datetime.datetime.strptime(date, '%m.%d.%y')
            sample_dates.append(sample_datetime)
          except Exception as e:
            logger.exception(e)
    else:
      for cell_ndx, cell in enumerate(col):
        #First cell is the Sample Site name
        if cell_ndx == 0:
          site_name = cell.value
          if site_name not in sample_data:
            sample_data[site_name] = []
        #Subsequent cells are the sample values.
        else:
          sample_value = cell.value
          sample_data[site_name].append({
            'site_name': site_name,
            'sample_date': sample_dates[cell_ndx-1],
            'value': sample_value
          })


  logger.debug("Finished parse_sheet_data")

  return sample_data, sample_dates

def parse_all_beach_sheet_data(xl_file_name, sites_aliases):
  logger = logging.getLogger('check_email_for_update_logger')
  logger.debug("Starting parse_sheet_data, parsing file: %s" % (xl_file_name))


  sample_data = {}
  wb = load_workbook(filename = xl_file_name)
  bacteria_data_sheet = wb['AllBeachData_Admin']
  date_column = 'E'
  time_column = 'F'
  station_column = 'D'
  entero_column = 'G'
  sample_dates = []
  for row_num in range(2,bacteria_data_sheet.max_row):
    cell_name = "%s%d" % (station_column, row_num)
    cell_value = "%s%d" % (entero_column, row_num)
    cell_date = "%s%d" % (date_column, row_num)
    cell_time = "%s%d" % (time_column, row_num)
    try:
      if isinstance(bacteria_data_sheet[cell_time].value, datetime.time):
        sample_date_time = datetime.datetime.combine(bacteria_data_sheet[cell_date].value, bacteria_data_sheet[cell_time].value)
      else:
        sample_date_time = bacteria_data_sheet[cell_date].value
    except Exception as e:
      sample_date_time = bacteria_data_sheet[cell_date].value
      logger.exception(e)

    site_name = bacteria_data_sheet[cell_name].value.strip().lower()
    if site_name.lower() in sites_aliases:
      site_name = sites_aliases[site_name.lower()]

    if site_name not in sample_data:
      sample_data[site_name] = []
    sample_data[site_name].append({
      'site_name': site_name,
      'sample_date': sample_date_time,
      'value': bacteria_data_sheet[cell_value].value
    })

    #sample_data[site_name].append({
    #  'sample_date': sample_date_time,
    #  'value': bacteria_data_sheet[cell_value].value
    #})
    if bacteria_data_sheet[cell_date].value not in sample_dates:
      sample_dates.append(bacteria_data_sheet[cell_date].value)
    #if sample_date is None:
    #  sample_date = bacteria_data_sheet[cell_date].value
  if logger:
    logger.debug("Finished parse_sheet_data")

  return sample_data, sample_dates

def build_feature(site, sample_date, values, logger):

  logger = logging.getLogger('build_feature_logger')
  logger.debug("Starting build_feature_logger")
  logger.debug("Adding feature site: %s Desc: %s" % (site.name, site.description))

  beachadvisories = {
    'date': '',
    'station': site.name,
    'value': ''
  }
  if len(values):
    beachadvisories = {
      'date': sample_date,
      'station': site.name,
      'value': values
    }
  feature = {
    'type': 'Feature',
    'geometry': {
      'type': 'Point',
      'coordinates': [site.object_geometry.x, site.object_geometry.y]
    },
    'properties' : {
      'locale': site.description,
      'sign': False,
      'station': site.name,
      'epaid': site.epa_id,
      'beach': site.county,
      'desc': site.description,
      'len': '',
      'test': {
        'beachadvisories': beachadvisories
      }
    }
  }
  return feature

def build_current_file(data_dict, date_keys, config_file, fl_sites, build_missing):
  logger = logging.getLogger('build_current_file_logger')
  logger.debug("Starting build_current_file")

  try:
    advisory_results_filename = config_file.get('json_settings', 'advisory_results')
    station_results_directory = config_file.get('json_settings', 'station_results_directory')

  except ConfigParser.Error as e:
    if logger:
      logger.exception(e)
  else:
    features = []
    #Let's go site by site and find the data in the worksheet. Currently we only get the
    #data for Sarasota county.
    values = []
    bacteria_data = {}
    for site in fl_sites:
      #Find the most recent sample dates. Currently for Manatee we don't get the actual
      #bacteria sample data, so we harvest the advisory(either there is an advisory or not)
      #from the florida site: http://www.floridahealth.gov/environmental-health/beach-water-quality/county-detail.html?County=Manatee&Zip=34208-1928
      #The based on the site we get the data for the most current date, which may not
      #be the same for Manatee Country as it is for Sarasota.
      site_name = None
      for ndx, sample_date in reversed(list(enumerate(date_keys))):
        data = data_dict[sample_date]
        if site.description.lower() in data or site.name.lower() in data:
          bacteria_data = data
          break
      if logger:
        logger.debug("Site: %s Desc: %s searching data" % (site.name, site.description))
      values = None
      station_data = None
      if site.description.lower() in bacteria_data or site.name.lower() in bacteria_data:
        if logger:
          logger.debug("Adding feature site: %s Desc: %s" % (site.name, site.description))

        #Build the json structure the web app is going to use.
        try:
          station_data = bacteria_data[site.description.lower()]
          site_name = site.description.lower()
        except KeyError as e:
          try:
            station_data = bacteria_data[site.name.lower()]
            site_name = site.name.lower()
          except KeyError as e:
            continue
        if isinstance(station_data['value'], int) or isinstance(station_data['value'], long):
          values = str(station_data['value'])
        else:
          values = station_data['value'].split(';')
          #values = [int(val.strip()) for val in values]
          values = [val.strip() for val in values]
      elif build_missing:
        #values = None
        values = []

      if values is not None:
        if station_data is not None:
          feature = build_feature(site, station_data['sample_date'].strftime('%Y-%m-%d %H:%M:%S'), values, logger)
          features.append(feature)
        else:
          logger.error(f"No station data for site: {site.name}")
    try:
      with open(advisory_results_filename, "w") as json_out_file:
        #Now output JSON file.
        json_data = {
          'type': 'FeatureCollection',
          'features': features
        }
        json_out_file.write(json.dumps(json_data, sort_keys=True))
    except (json.JSONDecodeError, IOError) as e:
      if logger:
        logger.exception(e)

  if logger:
    logger.debug("Finished build_current_file")

  return

def build_station_file(bacteria_data, data_date, config_file, fl_sites, build_missing):
  logger = logging.getLogger('build_station_file_logger')
  logger.debug("Starting build_station_file")

  try:
    station_results_directory = config_file.get('json_settings', 'station_results_directory')

  except ConfigParser.Error as e:
    if logger:
      logger.exception(e)
  else:
    initial_data_length = len(bacteria_data)
    for site in fl_sites:
      if logger:
        logger.debug("Searching for site: %s in bacteria data." % (site.description.lower()))
      values = None
      if site.description.lower() in bacteria_data or site.name.lower() in bacteria_data:
        try:
          if initial_data_length != len(bacteria_data):
            logger.error(f"Initial length: {initial_data_length} does not match current length: {len(bacteria_data)}")
          station_data = bacteria_data[site.description.lower()]
        except KeyError as e:
          try:
            station_data = bacteria_data[site.name.lower()]
          except KeyError as e:
            logger.exception(e)
            continue


        if isinstance(station_data['value'], int):# or isinstance(station_data['value'], long):
          values = str(station_data['value'])
        else:
          values = station_data['value'].split(';')
          #values = [int(val.strip()) for val in values]
          values = [val.strip() for val in values]
      else:
        if build_missing:
          values = []
          logger.debug("Site: %s not found in bacteria data" % (site.description.lower()))
      if values is not None:
        if logger:
          logger.debug("Station: %s building file." % (site.description.lower()))
        feature = None
        #Now find if we have the station file already, if not we create it.
        station_filename = os.path.join(station_results_directory, '%s.json' % (site.name))
        logger.debug("Processing station file: %s" % (station_filename))
        test_data = None
        if len(values):
          test_data = {
            #'date': station_data['sample_date'].strftime('%Y-%m-%d %H:%M:%S'),
            'date': data_date.strftime('%Y-%m-%d %H:%M:%S'),
            'station': site.name,
            'value': values
          }
        #Do we have a station file already, if so get the data.
        if os.path.isfile(station_filename):
          if test_data is not None:
            try:
              logger.debug("Opening station JSON file: %s" % (station_filename))
              with open(station_filename, 'r') as station_json_file:
                feature = json.loads(station_json_file.read())
                if feature is not None:
                  if 'test' in feature['properties']:
                    beachadvisories = feature['properties']['test']['beachadvisories']
                  else:
                    beachadvisories = []
                  #Make sure the date is not already in the list.
                  if not contains(beachadvisories, lambda x: x['date'] == test_data['date']):
                    if logger:
                      logger.debug("Station: %s adding date: %s" % (site.name, test_data['date']))
                    beachadvisories.append(test_data)
                    beachadvisories.sort(key=lambda x: x['date'], reverse=False)
            except (json.JSONDecodeError, IOError, Exception) as e:
              if logger:
                logger.exception(e)
        #No file, so let's create the station data
        else:
          logger.debug("Creating station JSON file: %s" % (station_filename))
          beachadvisories = []
          if test_data is not None:
            beachadvisories = [test_data]
          feature = {
            'type': 'Feature',
            'geometry': {
              'type': 'Point',
              'coordinates': [site.object_geometry.x, site.object_geometry.y]
            },
            'properties' : {
              'locale': site.description,
              'sign': False,
              'station': site.name,
              'epaid': site.epa_id,
              'beach': site.county,
              'desc': site.description,
              'len': '',
              'test': {
                'beachadvisories': beachadvisories
              }
            }
          }
        try:
          if feature is not None:
            with open(station_filename, 'w') as station_json_file:
              feature_json = json.dumps(feature)
              logger.debug("Feature: %s" % (feature_json))
              station_json_file.write(feature_json)
          else:
            logger.debug("Feature is None")
        except (json.JSONDecodeError, IOError) as e:
          if logger:
            logger.exception(e)
      else:
        if logger:
          logger.debug("Site: %s not found in bacteria data." % (site.description.lower()))


  if logger:
    logger.debug("Finished build_station_file")

  return

def get_sampling_data(config_filename, data_dict, sites_aliases):
    config_file = ConfigParser.RawConfigParser()
    config_file.read(config_filename)

    logger = logging.getLogger()
    logger.debug("Starting get_sampling_data")

    swim_adv = florida_wq_advisory('sarasota', config_filename)
    data_file_list = check_email_for_update(config_file)
    logger.debug("Received %d wq file emails" % (len(data_file_list)))
    if len(data_file_list):
      for data_file in data_file_list:
        #For each file, we need to check what type it is. We get results from SUn Coast Water
        #keeper and Sarasota.
        wb = load_workbook(filename=data_file)
        #This is a XLS file from Sarasota.
        if 'SCHD Bact Results' in wb.sheetnames:
          process_sarasota_data(data_file, data_dict)
        elif 'AllBeachData_Admin' in wb.sheetnames:
          process_all_sample_data(data_file, data_dict, sites_aliases)
        #This is probably the newer Sun Coast spreadsheet
        elif 'Sheet1' in wb.sheetnames:
          process_suncoast_data(data_file, data_dict)

def process_suncoast_data(data_file, data_dict):
  logger = logging.getLogger()
  try:
    sample_data, sample_dates = parse_suncoast_sheet_data(data_file)
    # Some sites can be sampled on different days if they test high one day.
    sites = sample_data.keys()
    for site in sites:
      # Normally a site only has the one sample date, but in the case of
      # resampling there may be multiple.
      for data in sample_data[site]:
        if isinstance(data['sample_date'], datetime.date):
          sample_date = data['sample_date'].date()
          if sample_date not in data_dict:
            data_dict[sample_date] = {}
          date_recs = data_dict[sample_date]
          date_recs[site] = data
        else:
          logger.error("Site: %s Invalid date: %s" % (site, data['sample_date']))
  except Exception as e:
    logger.exception(e)


def process_sarasota_data(data_file, data_dict):
  logger = logging.getLogger()

  try:
    sample_data, sample_dates = parse_sheet_data(data_file)

    # Some sites can be sampled on different days if they test high one day.
    sites = sample_data.keys()
    for site in sites:
      # Normally a site only has the one sample date, but in the case of
      # resampling there may be multiple.
      for data in sample_data[site]:
        if isinstance(data['sample_date'], datetime.date):
          sample_date = data['sample_date'].date()
          if sample_date not in data_dict:
            data_dict[sample_date] = {}
          date_recs = data_dict[sample_date]
          date_recs[site] = data
        else:
          logger.error("Site: %s Invalid date: %s" % (site, data['sample_date']))
  except Exception as e:
    logger.exception(e)
  return

def process_all_sample_data(data_file, data_dict, sites_aliases):
  start_time = time.time()
  logger = logging.getLogger()
  logger.debug("process_all_sample_data started.")
  try:
    sample_data, sample_dates = parse_all_beach_sheet_data(data_file, sites_aliases)

    # Some sites can be sampled on different days if they test high one day.
    sites = sample_data.keys()
    for site in sites:
      # Normally a site only has the one sample date, but in the case of
      # resampling there may be multiple.
      for data in sample_data[site]:
        if isinstance(data['sample_date'], datetime.date):
          sample_date = data['sample_date'].date()
          if sample_date not in data_dict:
            data_dict[sample_date] = {}
          date_recs = data_dict[sample_date]
          date_recs[site] = data
        else:
          logger.error(f"Site: {site} Invalid date: {data['sample_date']}")
  except Exception as e:
    logger.exception(e)

  logger.debug(f"process_all_sample_data finished in {time.time()-start_time} seconds.")

  return

def get_sarasota_county_data(config_filename, data_dict):
  config_file = ConfigParser.RawConfigParser()
  config_file.read(config_filename)

  logger = logging.getLogger('get_sarasota_county_data_logger')
  logger.debug("Starting get_sarasota_county_data")

  swim_adv = florida_wq_advisory('sarasota', config_filename)
  #advisories = swim_adv.get_advisories()
  #advisories_date = advisories[0]['date']
  data_file_list = check_email_for_update(config_file)
  logger.debug("Received %d wq file emails" % (len(data_file_list)))
  if len(data_file_list):
    for data_file in data_file_list:
      #if data_file == "/Users/danramage/Documents/workspace/WaterQuality/Florida_Water_Quality/data/sampling_data/Weekly_Sarasota_Data/02222016.xlsx":
      #  i = 0
      sample_data, sample_dates = parse_sheet_data(data_file)

      #Some sites can be sampled on different days if they test high one day.
      sites = sample_data.keys()
      for site in sites:
        #Normally a site only has the one sample date, but in the case of
        #resampling there may be multiple.
        for data in sample_data[site]:
          if isinstance(data['sample_date'], datetime.date):
            sample_date = data['sample_date'].date()
            if sample_date not in data_dict:
              data_dict[sample_date] = {}
            date_recs = data_dict[sample_date]
            date_recs[site] = data
          else:
            logger.error("Site: %s Invalid date: %s" % (site, data['sample_date']))
def get_manatee_county_data(config_filename, fl_sites, data_dict):
  results = []
  logger = logging.getLogger('get_manatee_county_data_logger')
  logger.debug("Starting get_manatee_county_data")
  #For now, we do not have the manatee bacteria sample data results.
  #The best we can do is use the current results from the advisory
  #page.
  swim_adv = florida_wq_advisory('manatee', config_filename)
  advisories = swim_adv.get_advisories()
  if len(advisories):
    advisories_date = advisories[0]['date']
    sample_data = {}
    for advisory in advisories:
      #An advisory occurs when the etcoc value is 104 or greater, so if the
      #advisory data we downloaded has an advisory issued, the bacteria count
      #has to be at least 104.
      logger.debug("Processing site: %s Date: %s advisory: %d" % (advisory['site_name'], advisory['date'], advisory['advisory']))
      value = 0
      if advisory['advisory']:
        value = 104
      sample_data[advisory['site_name']] = {
        'sample_date': advisory['date'],
        'value': value
      }
      sample_date = advisory['date']
      if sample_date.date() not in data_dict:
        data_dict[sample_date.date()] = {}

      rec = data_dict[sample_date.date()]
      rec[advisory['site_name']] = {
        'sample_date': advisory['date'],
        'value': value
      }

    results.append({'sample_date': advisories_date,
                    'sample_data': sample_data})
  else:
    logger.error("No advisories returned.")
  return results
def main():
  parser = optparse.OptionParser()
  parser.add_option("--ConfigFile", dest="config_file", default=None,
                    help="INI Configuration file." )
  parser.add_option("--CreateCurrentSampleFile", dest="create_current", default=False,
                    help="If set, this will create the current the stations json file with the latest data.")
  parser.add_option("--CreateStationsFile", dest="create_stations", default=False,
                    help="If set, this will create or update the individual station json file with the data from latest email.")
  parser.add_option("--DateFile", dest="date_file", default=None,
                    help="If provided, full file path for a list of dates processed.")
  parser.add_option("--SiteAliases", dest="site_aliases", default=None,
                    help="If provided, full file path for a list of dates processed.")

  (options, args) = parser.parse_args()

  if options.config_file is None:
    parser.print_help()
    sys.exit(-1)

  try:
    config_file = ConfigParser.RawConfigParser()
    config_file.read(options.config_file)
  except Exception as e:
    raise
  else:
    logger = None
    try:
      logConfFile = config_file.get('logging', 'config_file')
      boundaries_location_file = config_file.get('boundaries_settings', 'boundaries_file')
      sites_location_file = config_file.get('boundaries_settings', 'sample_sites')
      site_aliases = {}
      if options.site_aliases is not None:
        with open(options.site_aliases, "r") as fobj_site_aliases:
          sites_aliases = json.load(fobj_site_aliases)
      if logConfFile:
        logging.config.fileConfig(logConfFile)
        logger = logging.getLogger('florida_wq_bact_harvest_logger')
        logger.info("Log file opened.")
    except ConfigParser.Error as e:
      if logger:
        logger.exception(e)
    else:

      fl_sites = florida_sample_sites(True)
      fl_sites.load_sites(file_name=sites_location_file, boundary_file=boundaries_location_file)

      data_dict = {}
      results = get_sampling_data(options.config_file, data_dict, sites_aliases)
      #sarasota_results = get_sarasota_county_data(options.config_file, data_dict)
      #manatee_results = get_manatee_county_data(options.config_file, fl_sites, data_dict)

      date_keys = data_dict.keys()
      sorted(date_keys)
      #Build the individual station json files.
      for date_key in date_keys:
        build_station_file(data_dict[date_key], date_key, config_file, fl_sites, True)
      #Build the most current results for all the stations.
      build_current_file(data_dict, date_keys, config_file, fl_sites, True)

      if options.date_file is not None:
        with open(options.date_file, 'w') as date_file_obj:
          if date_file_obj is not None:
            for date_key in date_keys:
              date_file_obj.write('%s,' % (date_key))
            date_file_obj.write('\n')

  return


if __name__ == "__main__":
  main()
