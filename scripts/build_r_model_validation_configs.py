import sys
sys.path.append('../commonfiles')
import os
import logging.config

from openpyxl import load_workbook
import poplib
import email
import optparse
import ConfigParser
import simplejson as json
import datetime
import traceback
from collections import OrderedDict

from florida_wq_data import florida_sample_sites

VB_FUNCTIONS_REPLACE = [
  ("POLY", "VB_POLY"),
  ("QUADROOT", "VB_QUADROOT"),
  ("SQUAREROOT", "VB_SQUAREROOT"),
  ("INVERSE", "VB_INVERSE"),
  ("SQUARE", "VB_SQUARE"),
  ("WindO_comp", "VB_WindO_comp"),
  ("WindA_comp", "VB_WindA_comp"),
  ("LOG10", "VB_LOG10")
]

def read_sheet_data(work_sheet, site_name, output_dir):
  sheet_data = []
  #First two columns should always be the autonumber and Log10(ent)
  col_count = work_sheet.max_column
  row_count = work_sheet.max_row
  output_filename = os.path.join(output_dir, "%s%s.csv" % (site_name, work_sheet.title))
  with open(output_filename, 'w') as csv_output_file:
    for row_num in range(1, row_count+1):
      row_data = []
      print "Processing row: %d" % (row_num)
      #Start at column 2 since we don't need the first column.
      for col_num in range(2, col_count+1):
        print "Processing col: %d" % (col_num)
        cell = work_sheet.cell(row=row_num, column=col_num)
        #Header row
        if row_num == 1:
          if cell.value is not None:
            if cell.value not in sheet_data:
              col_name = cell.value.strip()
              if col_name == 'Log10(ent)':
                col_name = 'Log10ent'
              sheet_data.append(col_name)
            else:
              print("ERROR: Duplicate column name")
              sys.exit(-1)
        if cell.value is not None:
          value = cell.value
          if row_num == 1:
            value = value.strip()
            if value == 'Log10(ent)':
              value = 'Log10ent'
            else:
              for vb_replacement in VB_FUNCTIONS_REPLACE:
                if cell.value.find(vb_replacement[0]) != -1:
                  value = cell.value.replace('[', '_').replace(']', '_').replace(',', '_')
                  break

          row_data.append(str(value))
      csv_output_file.write(",".join(row_data))
      csv_output_file.write("\n")
  return sheet_data

def write_ini_file(site, models, config_file):
  try:
    config_file_outdir = config_file.get("r_settings", "config_file_directory")
    figure_output_directory = config_file.get("r_settings", "figure_output_directory")
    model_results_output_directory = config_file.get("r_settings", "model_results_output_directory")
    model_data_directory = config_file.get("r_settings", "model_data_directory")

    model_config_parser = ConfigParser.ConfigParser()
    model_config_outfile_name = os.path.join(config_file_outdir, '%s.ini' % (site.name))
  except ConfigParser.Error, e:
    traceback.print_exc(e)
  else:
    model_count = len(models)
    model_config_parser.add_section("settings")
    model_config_parser.set("settings", "site_name", site.name)
    model_config_parser.set("settings", "model_count", model_count)
    model_config_parser.set("settings", "figure_output_directory", figure_output_directory)
    model_config_parser.set("settings", "model_results_output_directory", model_results_output_directory)
    model_cnt = 1
    for model in models:
      model_section = "model_%d" % (model_cnt)
      print "Model: %s" % (model_section)
      model_config_parser.add_section(model_section)
      data_file = os.path.join(model_data_directory, '%s%s.csv' % (site.name, model))
      model_config_parser.set(model_section, "model_name", model)
      model_config_parser.set(model_section, "input_data", data_file)
      model_config_parser.set(model_section, "roc_image_title", '%s %s Model ROC Curve' % (site.name, model))
      model_config_parser.set(model_section, "performance_indicators_title", '%s Summary' % (model))
      #modelA1<-lm(datA1$Log10ent ~ datA1$c10_min_salinity_144, data=datA1, na.action=na.omit)
      model_data_cols = []
      for data_col in models[model]:
        for vb_replacement in VB_FUNCTIONS_REPLACE:
          if data_col.find(vb_replacement[0]) != -1:
            data_col= data_col.replace('[', '_').replace(']', '_').replace(',', '_')
            break
        model_data_cols.append(data_col)

      data_cols = [model_data_cols[ndx] for ndx in range(1, len(model_data_cols))]
      linear_model_str = "lm(%s ~ %s, data=modelDat, na.action=na.omit)"\
                         % (model_data_cols[0],
                            "+".join(data_cols))
      model_config_parser.set(model_section, "linear_model", linear_model_str)
      model_cnt += 1
  try:
    with open(model_config_outfile_name, 'w') as model_config_ini_obj:
      model_config_parser.write(model_config_ini_obj)
  except IOError, e:
    traceback.print_exc(e)

  return
def main():
  parser = optparse.OptionParser()
  parser.add_option("-c", "--ConfigFile", dest="config_file", default=None,
                    help="INI Configuration file." )
  parser.add_option("-s", "--SiteName", dest="site_name", default=None,
                    help="Site Name" )
  parser.add_option("-x", "--SiteXLDataFile", dest="site_xl_file", default=None,
                    help="Excel Data File" )

  (options, args) = parser.parse_args()

  if options.config_file is None:
    parser.print_help()
    sys.exit(-1)

  try:
    config_file = ConfigParser.RawConfigParser()
    config_file.read(options.config_file)
  except Exception, e:
    raise
  else:
    try:

      boundaries_location_file = config_file.get('boundaries_settings', 'boundaries_file')
      sites_location_file = config_file.get('boundaries_settings', 'sample_sites')
      station_model_dir = config_file.get('station_model_files', 'prediction_config_dir')
      config_file_directory = config_file.get('r_settings', 'config_file_directory')
    except ConfigParser.Error, e:
      traceback.print_exc(e)
    else:
      fl_sites = florida_sample_sites(True)
      fl_sites.load_sites(file_name=sites_location_file, boundary_file=boundaries_location_file)

      for site in fl_sites:
        if site.name == options.site_name:
          site_data_wb = load_workbook(filename = options.site_xl_file)
          site_config_filename = os.path.join(station_model_dir, '%s.ini' % (site.name))
          try:
            site_config_file = ConfigParser.RawConfigParser()
            site_config_file.read(site_config_filename)
          except Exception, e:
            traceback.print_exc(e)
          else:
            models = OrderedDict()
            sheet_names = site_data_wb.get_sheet_names()
            for ws_name in sheet_names:
              work_sheet = site_data_wb[ws_name]
              print "Processing sheet: %s" % (ws_name)
              wb_data = read_sheet_data(work_sheet, site.name, config_file_directory)
              models[ws_name] = wb_data
            write_ini_file(site, models, config_file)
  return


if __name__ == "__main__":
  main()
